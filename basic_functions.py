from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.alert import Alert
import openpyxl


def open_browser_to_link(url):
    service = Service("C:/Users/User/OneDrive/Desktop/Hiring_Selenium/Constants/driver/chromedriver.exe")
    driver = webdriver.Chrome(service=service)
    driver.maximize_window()
    driver.get(url)
    return driver
 
    
def click_element(driver, xpath):
    wait = WebDriverWait(driver, 10) 
    element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    element.click()
 
   
def enter_text_in_element(driver, xpath, text):
    wait = WebDriverWait(driver, 10) 
    element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    element.send_keys(text)


def get_element_text(driver, xpath):
    wait = WebDriverWait(driver, 10) 
    element = wait.until(EC._element_if_visible((By.XPATH, xpath)))
    return element.text

def scroll_element_into_view(driver, xpath):
    wait =  WebDriverWait(driver, 10)
    # element = wait.until(EC._element_if_visible((By.XPATH, xpath)))
    element = driver.find_element(By.XPATH, xpath)
    driver.execute_script("arguments[0].scrollIntoView();", element)

def select_value_in_dropdown(driver, xpath, value):
    wait = WebDriverWait(driver, 10)  
    element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    dropdown = Select(element)
    dropdown.select_by_value(value) 

def handle_alert(driver, accept=True):
    alert = Alert(driver)
    alert_text_message = alert.text
    
    if accept:
        alert.accept()
    else :
        alert.dismiss()

    return alert_text_message
    # wait = WebDriverWait(driver, 10)    
    
def wait_until_element_is_visible(driver, xpath, wait_time=10):
    wait = WebDriverWait(driver, 10)
    try:
        element = wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))
        if element:
            return f"Element Found: {xpath},"
    except Exception as e:
        print(f"{e}: Element not found: {xpath}.")
        
        
        
        
        
        
        
################################################## EXCEL UTILITIES #####################################################


def getRowCount (file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount (file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData (file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
